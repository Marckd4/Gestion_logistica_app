
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

from bodegabsf.forms import BsfForm
from .models import Bsf  # importar el modelo

def data(request):
    bsfs = Bsf.objects.all()
    return render(request, 'index.html', context={'bsfs': bsfs})

# area formulario 

def formulario(request):
    if request.method == 'POST':
        form = BsfForm(request.POST)
        if form.is_valid():
             form.save()
             return HttpResponseRedirect('/bodegabsf')
    else:
        form = BsfForm()
        
        
    return render( request, 'bsf_form.html',{'form': form})



# area de exportar excel 
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Bsf

def limpiar(valor):
    if valor is None:
        return ""
    return str(valor)

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    columnas = [
        "Categoria", "Empresa", "Ubicacion", "Cod_Ean", "Cod_Dun", "Cod_Sistema",
        "Descripcion", "Unidad", "Pack", "FactorX", "Cajas", "Saldo",
        "Stock_Fisico", "Observacion", "Fecha_Venc", "Fecha_Imp",
        "Fecha_Inv", "Encargado"
    ]

    # Encabezados
    for col_num, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=titulo)
        cell.font = Font(bold=True)

    data = Bsf.objects.all()

    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=limpiar(item.categoria))
        ws.cell(row=row_num, column=2, value=limpiar(item.empresa))
        ws.cell(row=row_num, column=3, value=limpiar(item.ubicacion))
        ws.cell(row=row_num, column=4, value=limpiar(item.cod_ean))
        ws.cell(row=row_num, column=5, value=limpiar(item.cod_dun))
        ws.cell(row=row_num, column=6, value=limpiar(item.cod_sistema))
        ws.cell(row=row_num, column=7, value=limpiar(item.descripcion))
        ws.cell(row=row_num, column=8, value=limpiar(item.unidad))
        ws.cell(row=row_num, column=9, value=limpiar(item.pack))
        ws.cell(row=row_num, column=10, value=limpiar(item.factorx))
        ws.cell(row=row_num, column=11, value=limpiar(item.cajas))
        ws.cell(row=row_num, column=12, value=limpiar(item.saldo))
        ws.cell(row=row_num, column=13, value=limpiar(item.stock_fisico))
        ws.cell(row=row_num, column=14, value=limpiar(item.observacion))
        ws.cell(row=row_num, column=15, value=limpiar(item.fecha_venc))
        ws.cell(row=row_num, column=16, value=limpiar(item.fecha_imp))
        ws.cell(row=row_num, column=17, value=limpiar(item.fecha_inv))
        ws.cell(row=row_num, column=18, value=limpiar(item.encargado))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=\"inventario_bodega_BSF.xlsx\"'

    wb.save(response)
    return response


# eliminar y editar tabla 


from django.shortcuts import render, redirect, get_object_or_404
from .models import Bsf
from .forms import BsfForm

# ---------- EDITAR ----------
def editar_bsf(request, id):
    bsf = get_object_or_404(Bsf, id=id)

    if request.method == "POST":
        form = BsfForm(request.POST, instance=bsf)
        if form.is_valid():
            form.save()
            return redirect('data')  # ← Ajusta al nombre de tu vista principal
    else:
        form = BsfForm(instance=bsf)

    return render(request, 'editar_bsf.html', {'form': form, 'bsf': bsf})


# ---------- ELIMINAR ----------
def eliminar_bsf(request, id):
    bsf = get_object_or_404(Bsf, id=id)

    if request.method == "POST":
        bsf.delete()
        return redirect('data')   # ← Ajusta al nombre de tu vista principal

    return render(request, 'eliminar_bsf.html', {'bsf': bsf})



# autocompletar del dun en formulario 

from django.shortcuts import render
from django.http import JsonResponse
from .models import Bsf
from .forms import BsfForm

def formulario(request, id=None):
    mensaje = ""
    if id:  # modo edición
        instancia = Bsf.objects.get(id=id)
    else:
        instancia = None

    if request.method == "POST":
        form = BsfForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            mensaje = "✅ Producto guardado correctamente."
    else:
        form = BsfForm(instance=instancia)

    return render(request, 'bsf_form.html', {"form": form, "mensaje": mensaje})

def buscar_producto(request):
    cod_dun = request.GET.get('cod_dun', '').strip()
    resultados = []

    if cod_dun:
        productos = Bsf.objects.filter(cod_dun__icontains=cod_dun)[:10]  # máximo 10 coincidencias
        for p in productos:
            resultados.append({
                "cod_dun": p.cod_dun,
                "cod_ean": p.cod_ean,
                "cod_sistema": p.cod_sistema,
                "descripcion": p.descripcion,
            })

    return JsonResponse({"resultados": resultados})
