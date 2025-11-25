from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

from bodegacentral.forms import CentralForm
from .models import Central # IMPORTAR EL MODEL 

def index(request):
    centrales = Central.objects.all()
   
    return render(request, 'central.html', context={'centrales':centrales})

# area formulario 

def formulario(request):
    if request.method == 'POST':
        form = CentralForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/bodegacentral')
    else:
        form = CentralForm()
            
    
    return render( request, 'central_form.html',{'form': form})



# exportar excel 

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse


def limpiar_valor(valor):
    return "" if valor in (None, "", 0) else valor

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Columnas EXACTAS igual a la tabla HTML
    columnas = [
        "Categoria", "Empresa", "Ubicacion", "Cod_Ean", "Cod_Dun", "Cod_Sistema",
        "Descripcion", "Unidad", "Pack", "FactorX", "Cajas", "Saldo",
        "Stock_Fisico", "Observacion", "Fecha_Venc", "Fecha_Imp",
        "Contenedor", "Fecha_Inv", "Encargado",
    ]

    # Escribir encabezados
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    data = Central.objects.all()

    # Escribir registros
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=limpiar_valor(item.categoria))
        ws.cell(row=row_num, column=2, value=limpiar_valor(item.empresa))
        ws.cell(row=row_num, column=3, value=limpiar_valor(item.ubicacion))
        ws.cell(row=row_num, column=4, value=limpiar_valor(item.cod_ean))
        ws.cell(row=row_num, column=5, value=limpiar_valor(item.cod_dun))
        ws.cell(row=row_num, column=6, value=limpiar_valor(item.cod_sistema))
        ws.cell(row=row_num, column=7, value=limpiar_valor(item.descripcion))
        ws.cell(row=row_num, column=8, value=limpiar_valor(item.unidad))
        ws.cell(row=row_num, column=9, value=limpiar_valor(item.pack))
        ws.cell(row=row_num, column=10, value=limpiar_valor(item.factorx))
        ws.cell(row=row_num, column=11, value=limpiar_valor(item.cajas))
        ws.cell(row=row_num, column=12, value=limpiar_valor(item.saldo))
        ws.cell(row=row_num, column=13, value=limpiar_valor(item.stock_fisico))
        ws.cell(row=row_num, column=14, value=limpiar_valor(item.observacion))
        ws.cell(row=row_num, column=15, value=limpiar_valor(item.fecha_venc))
        ws.cell(row=row_num, column=16, value=limpiar_valor(item.fecha_imp))
        ws.cell(row=row_num, column=18, value=limpiar_valor(item.fecha_inv))
        ws.cell(row=row_num, column=19, value=limpiar_valor(item.encargado))

    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inventario_Bodega_Lingues.xlsx"'

    wb.save(response)
    return response



#eliminar - editar

from django.shortcuts import render, redirect, get_object_or_404
from .models import Central
from .forms import CentralForm

# -------- EDITAR CENTRAL --------
def editar_central(request, id):
    central = get_object_or_404(Central, id=id)

    if request.method == "POST":
        form = CentralForm(request.POST, instance=central)
        if form.is_valid():
            form.save()
            return redirect('index')   # ← ajusta al nombre de tu vista principal
    else:
        form = CentralForm(instance=central)

    return render(request, 'editar_central.html', {'form': form, 'central': central})


# -------- ELIMINAR CENTRAL --------
def eliminar_central(request, id):
    central = get_object_or_404(Central, id=id)

    if request.method == "POST":
        central.delete()
        return redirect('index')   # ← ajusta al nombre real de tu URL

    return render(request, 'eliminar_central.html', {'central': central})


