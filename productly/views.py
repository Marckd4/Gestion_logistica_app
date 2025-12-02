
from django.shortcuts import render
from bodegabsf.models import Bsf
from bodegacentral.models import Central
from collections import defaultdict

def inicio(request):
    return render(request, 'inicio.html')

# area resumen datos 

def resumen_unificado(request):

    resumen_dict = defaultdict(lambda: {
        "cod_dun": "",
        "cod_ean": "",
        "cod_sistema_bsf": "",
        "cod_sistema_central": "",
        "descripcion": "",
        "cajas_bsf": 0,
        "cajas_central": 0,
        "stock_bsf": 0,
        "stock_central": 0,
    })

    # ---- BSF (clave = DUN + Sistema) ----
    for b in Bsf.objects.all():
        key = f"{b.cod_dun}_{b.cod_sistema}"
        d = resumen_dict[key]

        d["cod_dun"] = b.cod_dun
        d["cod_ean"] = b.cod_ean
        d["cod_sistema_bsf"] = b.cod_sistema
        d["descripcion"] = b.descripcion
        d["cajas_bsf"] += b.cajas or 0
        d["stock_bsf"] += b.stock_fisico or 0

    # ---- CENTRAL (clave = DUN + Sistema) ----
    for c in Central.objects.all():
        key = f"{c.cod_dun}_{c.cod_sistema}"
        d = resumen_dict[key]

        d["cod_dun"] = c.cod_dun
        d["cod_ean"] = c.cod_ean
        d["cod_sistema_central"] = c.cod_sistema
        d["descripcion"] = c.descripcion
        d["cajas_central"] += c.cajas or 0
        d["stock_central"] += c.stock_fisico or 0

    # Convertimos a lista
    resumen = []
    for d in resumen_dict.values():
        d["total_cajas"] = d["cajas_bsf"] + d["cajas_central"]
        resumen.append(d)

    return render(request, "resumen_unificado.html", {"resumen": resumen})




def exportar_resumen_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse
    from collections import defaultdict
    from bodegabsf.models import Bsf
    from bodegacentral.models import Central

    # ==== Construcci칩n del resumen unificado ====
    resumen_dict = defaultdict(lambda: {
        "cod_dun": "",
        "cod_ean": "",
        "cod_sistema_bsf": "",
        "cod_sistema_central": "",
        "descripcion": "",
        "cajas_bsf": 0,
        "cajas_central": 0,
        "stock_bsf": 0,
        "stock_central": 0,
    })

    # ---- Datos BSF ----
    for b in Bsf.objects.all():
        key = f"{b.cod_dun}_{b.cod_sistema}"
        d = resumen_dict[key]
        d["cod_dun"] = b.cod_dun
        d["cod_ean"] = b.cod_ean
        d["cod_sistema_bsf"] = b.cod_sistema
        d["descripcion"] = b.descripcion
        d["cajas_bsf"] += b.cajas or 0
        d["stock_bsf"] += b.stock_fisico or 0

    # ---- Datos Central ----
    for c in Central.objects.all():
        key = f"{c.cod_dun}_{c.cod_sistema}"
        d = resumen_dict[key]
        d["cod_dun"] = c.cod_dun
        d["cod_ean"] = c.cod_ean
        d["cod_sistema_central"] = c.cod_sistema
        d["descripcion"] = c.descripcion
        d["cajas_central"] += c.cajas or 0
        d["stock_central"] += c.stock_fisico or 0

    # Convertir a lista y ordenar
    resumen = []
    for data in resumen_dict.values():
        data["total_cajas"] = data["cajas_bsf"] + data["cajas_central"]
        resumen.append(data)
    resumen = sorted(resumen, key=lambda x: x["cod_dun"])

    # ==== Crear Excel ====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Inventario"

    columnas = [
        "Cod DUN", "Cod EAN", "Cod Sistema BSF", "Cod Sistema Central", "Descripci칩n",
        "Cajas BSF", "Cajas Central", "Stock BSF", "Stock Central", "Total Cajas"
    ]

    # Encabezados
    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna).font = Font(bold=True)

    # Filas
    for row_num, item in enumerate(resumen, 2):
        ws.cell(row=row_num, column=1, value=item["cod_dun"])
        ws.cell(row=row_num, column=2, value=item["cod_ean"])
        ws.cell(row=row_num, column=3, value=item["cod_sistema_bsf"])
        ws.cell(row=row_num, column=4, value=item["cod_sistema_central"])
        ws.cell(row=row_num, column=5, value=item["descripcion"])
        for col, key in zip(range(6, 11), ["cajas_bsf","cajas_central","stock_bsf","stock_central","total_cajas"]):
            ws.cell(row=row_num, column=col, value=item[key]).alignment = Alignment(horizontal="right")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="resumen_unificado.xlsx"'
    wb.save(response)
    return response


# usuario

from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages

def crear_usuario(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if User.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya existe")
            return redirect("crear_usuario")

        # Crear usuario con todos los permisos
        User.objects.create_superuser(username=username, password=password)

        messages.success(request, "Usuario administrador creado correctamente")
        return redirect("crear_usuario")

    return render(request, "crear_usuario.html")

# dashboard

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def dashboard(request):
    return render(request, "dashboard.html")



# login usuario

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

def login_usuario(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('inicio')  # Ajusta si quieres otra p치gina de inicio
        else:
            messages.error(request, "Usuario o contrase침a incorrectos")

    return render(request, "login.html")
